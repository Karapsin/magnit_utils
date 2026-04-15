from __future__ import annotations

import re
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import load_workbook

_INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]\*:/\\?]")
_DEFAULT_SHEET_NAME = "Sheet1"
_MELTED_VALUE_COLUMN = "__pivot_and_break_value__"


def pivot_and_break_table(
    df: pd.DataFrame | Sequence[pd.DataFrame],
    rows: str,
    output: str | Path,
    value: str | Sequence[str] | None = None,
    columns: str | None = None,
    break_by: str | None = None,
    sheet_by: str | None = None,
    append: bool = False,
    enforce_same_row_order: bool = False,
) -> dict[object | None, list[pd.DataFrame]] | dict[object | None, list[list[pd.DataFrame]]]:
    """Pivot a long-format dataframe into Excel tables split by tables and sheets.

    Returns the written tables grouped by the original ``sheet_by`` values. When
    ``sheet_by`` is omitted, the result contains a single ``None`` key.
    """
    dataframes = _normalize_dataframe_inputs(df)
    sheet_table_groups: list[dict[object | None, list[tuple[object | None, pd.DataFrame]]]] = []
    for part_df in dataframes:
        value_columns = _normalize_value_columns(
            df=part_df,
            value=value,
            rows=rows,
            columns=columns,
            break_by=break_by,
            sheet_by=sheet_by,
        )
        normalized_df, normalized_value = _prepare_pivot_source(
            df=part_df,
            rows=rows,
            value_columns=value_columns,
            columns=columns,
            break_by=break_by,
            sheet_by=sheet_by,
        )

        sheet_table_groups.append(
            _build_sheet_tables(
                df=normalized_df,
                break_by=break_by,
                sheet_by=sheet_by,
                table_builder=lambda part, normalized_value=normalized_value: _build_pivot_table(
                    df=part,
                    rows=rows,
                    value=normalized_value,
                    columns=columns,
                ),
            )
        )
    if enforce_same_row_order and len(sheet_table_groups) > 1:
        _enforce_same_row_order_across_groups(
            sheet_table_groups=sheet_table_groups,
            rows=rows,
        )
    _write_tables(
        sheet_table_groups=sheet_table_groups,
        output=Path(output),
        break_by=break_by,
        sheet_by=sheet_by,
        append=append,
    )
    return _extract_written_tables(sheet_table_groups)


def break_table(
    df: pd.DataFrame | Sequence[pd.DataFrame],
    output: str | Path,
    break_by: str | None = None,
    sheet_by: str | None = None,
    append: bool = False,
) -> dict[object | None, list[pd.DataFrame]] | dict[object | None, list[list[pd.DataFrame]]]:
    """Write grouped dataframe slices as stacked tables across Excel sheets."""
    dataframes = _normalize_dataframe_inputs(df)
    sheet_table_groups: list[dict[object | None, list[tuple[object | None, pd.DataFrame]]]] = []
    for part_df in dataframes:
        _validate_break_input(df=part_df, break_by=break_by, sheet_by=sheet_by)
        sheet_table_groups.append(
            _build_sheet_tables(
                df=part_df,
                break_by=break_by,
                sheet_by=sheet_by,
                table_builder=lambda part: _build_raw_table(
                    part,
                    break_by=break_by,
                    sheet_by=sheet_by,
                ),
            )
        )
    _write_tables(
        sheet_table_groups=sheet_table_groups,
        output=Path(output),
        break_by=break_by,
        sheet_by=sheet_by,
        append=append,
    )
    return _extract_written_tables(sheet_table_groups)


def _normalize_dataframe_inputs(
    df: pd.DataFrame | Sequence[pd.DataFrame],
) -> list[pd.DataFrame]:
    if isinstance(df, pd.DataFrame):
        return [df]
    if not isinstance(df, Sequence) or isinstance(df, (str, bytes)):
        raise TypeError("'df' must be a dataframe or a sequence of dataframes.")

    dataframes = list(df)
    if not dataframes:
        raise ValueError("'df' must contain at least one dataframe.")
    if any(not isinstance(part, pd.DataFrame) for part in dataframes):
        raise TypeError("'df' must be a dataframe or a sequence of dataframes.")
    return dataframes


def _extract_written_tables(
    sheet_table_groups: list[dict[object | None, list[tuple[object | None, pd.DataFrame]]]],
) -> dict[object | None, list[pd.DataFrame]] | dict[object | None, list[list[pd.DataFrame]]]:
    sheet_values = _collect_sheet_values(sheet_table_groups)
    if len(sheet_table_groups) == 1:
        sheet_tables = sheet_table_groups[0]
        return {
            sheet_value: [table for _, table in sheet_tables.get(sheet_value, [])]
            for sheet_value in sheet_values
        }

    return {
        sheet_value: [
            [table for _, table in sheet_tables.get(sheet_value, [])]
            for sheet_tables in sheet_table_groups
        ]
        for sheet_value in sheet_values
    }


def _build_raw_table(
    df: pd.DataFrame,
    break_by: str | None,
    sheet_by: str | None,
) -> pd.DataFrame:
    drop_columns = [column for column in (break_by, sheet_by) if column is not None]
    return df.drop(columns=drop_columns, errors="ignore").reset_index(drop=True)


def _validate_pivot_input(
    df: pd.DataFrame,
    rows: str,
    value_columns: list[str],
    columns: str | None,
    break_by: str | None,
    sheet_by: str | None,
) -> None:
    required_columns = list(value_columns)
    for column in (columns, break_by, sheet_by):
        if column is not None:
            required_columns.append(column)

    if len(value_columns) == 1:
        required_columns.append(rows)

    missing = sorted(column for column in required_columns if column not in df.columns)
    if missing:
        raise ValueError(f"Input dataframe is missing required columns: {missing}.")

    role_columns = {
        "columns": columns,
        "break_by": break_by,
        "sheet_by": sheet_by,
    }
    if len(value_columns) == 1:
        role_columns["rows"] = rows
    seen: dict[str, str] = {}
    for role, column in role_columns.items():
        if column is None:
            continue
        if column in seen:
            raise ValueError(
                f"{role!r} and {seen[column]!r} must refer to different dataframe columns."
            )
        seen[column] = role

    if len(value_columns) == 1:
        _validate_pivot_uniqueness(
            df=df,
            rows=rows,
            columns=columns,
            break_by=break_by,
            sheet_by=sheet_by,
        )


def _normalize_value_columns(
    df: pd.DataFrame,
    value: str | Sequence[str] | None,
    rows: str,
    columns: str | None,
    break_by: str | None,
    sheet_by: str | None,
) -> list[str]:
    if value is None:
        reserved_columns = {rows}
        reserved_columns.update(column for column in (columns, break_by, sheet_by) if column is not None)
        value_columns = [column for column in df.columns if column not in reserved_columns]
    elif isinstance(value, str):
        value_columns = [value]
    else:
        value_columns = list(value)

    if not value_columns:
        raise ValueError("'value' must contain at least one dataframe column.")

    if any(not isinstance(column, str) for column in value_columns):
        raise ValueError("'value' must be a column name or a sequence of column names.")

    duplicates = sorted({column for column in value_columns if value_columns.count(column) > 1})
    if duplicates:
        raise ValueError(f"'value' contains duplicate columns: {duplicates}.")

    return value_columns


def _prepare_pivot_source(
    df: pd.DataFrame,
    rows: str,
    value_columns: list[str],
    columns: str | None,
    break_by: str | None,
    sheet_by: str | None,
) -> tuple[pd.DataFrame, str]:
    _validate_pivot_input(
        df=df,
        rows=rows,
        value_columns=value_columns,
        columns=columns,
        break_by=break_by,
        sheet_by=sheet_by,
    )

    if len(value_columns) == 1:
        return df, value_columns[0]

    protected_columns = {column for column in (columns, break_by, sheet_by) if column is not None}
    if rows in protected_columns:
        raise ValueError(
            f"'rows'={rows!r} conflicts with an existing grouping column when multiple value columns are provided."
        )
    if rows in df.columns and rows not in value_columns:
        raise ValueError(
            f"'rows'={rows!r} already exists in the dataframe; choose a different name for the melted metric column."
        )

    melted_df = df.melt(
        id_vars=[column for column in df.columns if column not in value_columns],
        value_vars=value_columns,
        var_name=rows,
        value_name=_MELTED_VALUE_COLUMN,
    )
    _validate_pivot_uniqueness(
        df=melted_df,
        rows=rows,
        columns=columns,
        break_by=break_by,
        sheet_by=sheet_by,
    )
    return melted_df, _MELTED_VALUE_COLUMN


def _validate_pivot_uniqueness(
    df: pd.DataFrame,
    rows: str,
    columns: str | None,
    break_by: str | None,
    sheet_by: str | None,
) -> None:
    uniqueness_columns = [column for column in (sheet_by, break_by, rows, columns) if column is not None]
    duplicate_mask = df.duplicated(subset=uniqueness_columns, keep=False)
    if duplicate_mask.any():
        duplicates = (
            df.loc[duplicate_mask, uniqueness_columns]
            .drop_duplicates()
            .to_dict(orient="records")
        )
        if columns is None:
            raise ValueError(
                f"Values are not unique for rows={rows!r} within the selected sheet/break groups: {duplicates}."
            )
        raise ValueError(
            f"Values are not unique for rows={rows!r} and columns={columns!r} within the selected sheet/break groups: {duplicates}."
        )


def _validate_break_input(
    df: pd.DataFrame,
    break_by: str | None,
    sheet_by: str | None,
) -> None:
    required_columns = [column for column in (break_by, sheet_by) if column is not None]
    missing = sorted(column for column in required_columns if column not in df.columns)
    if missing:
        raise ValueError(f"Input dataframe is missing required columns: {missing}.")

    if break_by is not None and break_by == sheet_by:
        raise ValueError("'break_by' and 'sheet_by' must refer to different dataframe columns.")


def _build_sheet_tables(
    df: pd.DataFrame,
    break_by: str | None,
    sheet_by: str | None,
    table_builder,
) -> dict[object | None, list[tuple[object | None, pd.DataFrame]]]:
    if sheet_by is None:
        return {
            None: _build_tables(
                df=df,
                break_by=break_by,
                table_builder=table_builder,
            )
        }

    sheet_tables: dict[object | None, list[tuple[object | None, pd.DataFrame]]] = {}
    for sheet_value, sheet_df in df.groupby(sheet_by, sort=False, dropna=False):
        sheet_tables[sheet_value] = _build_tables(
            df=sheet_df,
            break_by=break_by,
            table_builder=table_builder,
        )
    return sheet_tables


def _build_tables(
    df: pd.DataFrame,
    break_by: str | None,
    table_builder,
) -> list[tuple[object | None, pd.DataFrame]]:
    if break_by is None:
        return [(None, table_builder(df))]

    tables: list[tuple[object | None, pd.DataFrame]] = []
    for break_value, table_df in df.groupby(break_by, sort=False, dropna=False):
        table = table_builder(table_df)
        tables.append((break_value, table))
    return tables


def _build_pivot_table(
    df: pd.DataFrame,
    rows: str,
    value: str,
    columns: str | None,
) -> pd.DataFrame:
    if columns is None:
        return df[[rows, value]].reset_index(drop=True)

    pivot_df = df[[rows, columns, value]].copy()
    pivot_df[rows] = pd.Categorical(pivot_df[rows], categories=pd.unique(df[rows]), ordered=True)
    pivot_df[columns] = pd.Categorical(
        pivot_df[columns],
        categories=pd.unique(df[columns]),
        ordered=True,
    )

    table = pivot_df.pivot(index=rows, columns=columns, values=value)
    return table.reset_index().rename_axis(columns=None)


def _enforce_same_row_order_across_groups(
    sheet_table_groups: list[dict[object | None, list[tuple[object | None, pd.DataFrame]]]],
    rows: str,
) -> None:
    reference_group = sheet_table_groups[0]
    sheet_values = _collect_sheet_values(sheet_table_groups)
    for sheet_value in sheet_values:
        reference_tables = reference_group.get(sheet_value, [])
        for group_index, sheet_tables in enumerate(sheet_table_groups[1:], start=2):
            current_tables = sheet_tables.get(sheet_value, [])
            if len(current_tables) > len(reference_tables):
                raise ValueError(
                    f"Dataframe #{group_index} produced more tables than dataframe #1 for sheet {sheet_value!r}."
                )

            aligned_tables: list[tuple[object | None, pd.DataFrame]] = []
            for table_index, (break_value, table) in enumerate(current_tables):
                reference_break_value, reference_table = reference_tables[table_index]
                if not _sheet_value_equals(break_value, reference_break_value):
                    raise ValueError(
                        "Cannot enforce the same row order when grouped tables do not align by "
                        f"position for sheet {sheet_value!r}: dataframe #1 has break value "
                        f"{reference_break_value!r} and dataframe #{group_index} has {break_value!r} "
                        f"at table position {table_index + 1}."
                    )
                aligned_tables.append(
                    (
                        break_value,
                        _align_table_to_reference_rows(
                            table=table,
                            reference_table=reference_table,
                            rows=rows,
                            sheet_value=sheet_value,
                            break_value=break_value,
                            dataframe_index=group_index,
                            table_index=table_index,
                        ),
                    )
                )
            sheet_tables[sheet_value] = aligned_tables


def _align_table_to_reference_rows(
    table: pd.DataFrame,
    reference_table: pd.DataFrame,
    rows: str,
    sheet_value: object | None,
    break_value: object | None,
    dataframe_index: int,
    table_index: int,
) -> pd.DataFrame:
    reference_labels = list(reference_table[rows])
    row_labels = list(table[rows])
    extra_labels = [
        label for label in row_labels if not any(_sheet_value_equals(label, ref) for ref in reference_labels)
    ]
    if extra_labels:
        raise ValueError(
            f"Dataframe #{dataframe_index} contains extra row labels for sheet {sheet_value!r}, "
            f"break {break_value!r}, table position {table_index + 1}: {extra_labels}."
        )

    aligned_table = table.set_index(rows).reindex(reference_labels).reset_index()
    aligned_table.columns = table.columns
    return aligned_table


def _write_tables(
    sheet_table_groups: list[dict[object | None, list[tuple[object | None, pd.DataFrame]]]],
    output: Path,
    break_by: str | None,
    sheet_by: str | None,
    append: bool,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    existing_sheet_names: set[str] = set()

    if output.exists() and not append:
        output.unlink()

    if output.exists():
        workbook = load_workbook(output, read_only=True)
        try:
            existing_sheet_names = set(workbook.sheetnames)
        finally:
            workbook.close()
        mode = "a"
    else:
        mode = "w"

    sheet_values = _collect_sheet_values(sheet_table_groups)
    sheet_name_map = _build_sheet_name_map(
        sheet_values=sheet_values,
        sheet_by=sheet_by,
        existing_sheet_names=existing_sheet_names,
    )

    writer_kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "overlay"

    with pd.ExcelWriter(output, **writer_kwargs) as writer:
        for sheet_value in sheet_values:
            sheet_name = sheet_name_map[sheet_value]
            _write_sheet_blocks(
                writer=writer,
                sheet_name=sheet_name,
                sheet_table_groups=sheet_table_groups,
                sheet_value=sheet_value,
                break_by=break_by,
            )


def _collect_sheet_values(
    sheet_table_groups: list[dict[object | None, list[tuple[object | None, pd.DataFrame]]]],
) -> list[object | None]:
    sheet_values: list[object | None] = []
    for sheet_tables in sheet_table_groups:
        for sheet_value in sheet_tables:
            if any(_sheet_value_equals(sheet_value, existing) for existing in sheet_values):
                continue
            sheet_values.append(sheet_value)
    return sheet_values


def _sheet_value_equals(left: object | None, right: object | None) -> bool:
    if pd.isna(left) and pd.isna(right):
        return True
    return left == right


def _write_table_block(
    writer: pd.ExcelWriter,
    sheet_name: str,
    table: pd.DataFrame,
    break_value: object | None,
    break_by: str | None,
    startrow: int,
    startcol: int,
) -> None:
    if break_by is not None:
        title = pd.DataFrame({break_by: [break_value]})
        title.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=False,
            startrow=startrow,
            startcol=startcol,
        )
        startrow += 1

    table.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=startrow,
        startcol=startcol,
    )


def _write_sheet_blocks(
    writer: pd.ExcelWriter,
    sheet_name: str,
    sheet_table_groups: list[dict[object | None, list[tuple[object | None, pd.DataFrame]]]],
    sheet_value: object | None,
    break_by: str | None,
) -> None:
    tables_per_group = [sheet_tables.get(sheet_value, []) for sheet_tables in sheet_table_groups]
    startcols: list[int] = []
    current_startcol = 0
    for tables in tables_per_group:
        startcols.append(current_startcol)
        if tables:
            current_startcol += _table_group_width(tables=tables, break_by=break_by) + 1

    startrow = 0
    max_table_count = max((len(tables) for tables in tables_per_group), default=0)
    for table_index in range(max_table_count):
        row_height = 0
        for group_index, tables in enumerate(tables_per_group):
            if table_index >= len(tables):
                continue

            break_value, table = tables[table_index]
            _write_table_block(
                writer=writer,
                sheet_name=sheet_name,
                table=table,
                break_value=break_value,
                break_by=break_by,
                startrow=startrow,
                startcol=startcols[group_index],
            )
            row_height = max(row_height, _table_block_height(table=table, break_by=break_by))

        startrow += row_height


def _table_group_width(
    tables: list[tuple[object | None, pd.DataFrame]],
    break_by: str | None,
) -> int:
    return max((_table_block_width(table=table, break_by=break_by) for _, table in tables), default=0)


def _table_block_width(
    table: pd.DataFrame,
    break_by: str | None,
) -> int:
    return max(len(table.columns), 1 if break_by is not None else 0)


def _table_block_height(
    table: pd.DataFrame,
    break_by: str | None,
) -> int:
    return len(table.index) + 4 if break_by is not None else len(table.index) + 3


def _build_sheet_name_map(
    sheet_values: list[object | None],
    sheet_by: str | None,
    existing_sheet_names: set[str],
) -> dict[object | None, str]:
    used_names = set(existing_sheet_names)
    result: dict[object | None, str] = {}
    for sheet_value in sheet_values:
        base_name = _sanitize_sheet_name(sheet_value=sheet_value, sheet_by=sheet_by)
        result[sheet_value] = _deduplicate_sheet_name(base_name=base_name, used_names=used_names)
    return result


def _sanitize_sheet_name(sheet_value: object | None, sheet_by: str | None) -> str:
    if sheet_by is None:
        return _DEFAULT_SHEET_NAME

    if pd.isna(sheet_value):
        raw_name = f"{sheet_by}_NA"
    else:
        raw_name = str(sheet_value).strip()

    raw_name = raw_name.replace("\n", " ").replace("\r", " ")
    sanitized = _INVALID_SHEET_CHARS_RE.sub("_", raw_name).strip("' ")
    if not sanitized:
        sanitized = f"{sheet_by}_value"
    return sanitized[:31]


def _deduplicate_sheet_name(base_name: str, used_names: set[str]) -> str:
    candidate = base_name[:31] or _DEFAULT_SHEET_NAME
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    counter = 2
    while True:
        suffix = f" ({counter})"
        truncated = f"{base_name[: max(0, 31 - len(suffix))]}{suffix}" or _DEFAULT_SHEET_NAME
        if truncated not in used_names:
            used_names.add(truncated)
            return truncated
        counter += 1
