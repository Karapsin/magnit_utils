from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from analytics_toolkit.excel import break_table, pivot_and_break_table


def test_pivot_and_break_table_writes_multiple_sheets_and_tables(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users", "arpu", "arpu", "users", "users", "arpu", "arpu"],
            "ab_group": ["control", "test_1", "control", "test_1", "control", "test_1", "control", "test_1"],
            "qr_group": ["ALL", "ALL", "ALL", "ALL", "1", "1", "1", "1"],
            "start_dt": ["2026-03-30", "2026-03-30", "2026-03-30", "2026-03-30", "2026-04-01", "2026-04-01", "2026-04-01", "2026-04-01"],
            "value": [100, 110, 2.5, 2.7, 50, 55, 1.1, 1.2],
        }
    )

    output = tmp_path / "report.xlsx"
    tables = pivot_and_break_table(
        df=df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert list(tables) == ["2026-03-30", "2026-04-01"]
    assert len(tables["2026-03-30"]) == 1
    assert tables["2026-03-30"][0].columns.tolist() == ["metric", "control", "test_1"]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-03-30", "2026-04-01"]
        first_sheet = workbook["2026-03-30"]
        rows = list(first_sheet.iter_rows(values_only=True))
        assert rows[:4] == [
            ("ALL", None, None),
            ("metric", "control", "test_1"),
            ("users", 100, 110),
            ("arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_sanitizes_and_deduplicates_sheet_names(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users", "users", "users"],
            "ab_group": ["control", "test_1", "control", "test_1"],
            "sheet_bucket": [
                "Report/Name:One*?",
                "Report/Name:One*?",
                "X" * 40,
                "X" * 40 + " trailing",
            ],
            "value": [1, 2, 3, 4],
        }
    )

    output = tmp_path / "sanitized.xlsx"
    pivot_and_break_table(
        df=df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="sheet_bucket",
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [
            "Report_Name_One__",
            "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX",
            "XXXXXXXXXXXXXXXXXXXXXXXXXXX (2)",
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_rejects_duplicates_within_group_slices(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "control"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 120],
        }
    )

    with pytest.raises(ValueError, match="not unique"):
        pivot_and_break_table(
            df=df,
            rows="metric",
            value="value",
            output=tmp_path / "duplicates.xlsx",
            columns="ab_group",
            sheet_by="start_dt",
        )


def test_pivot_and_break_table_accepts_multiple_value_columns(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "start_dt": ["2026-03-30", "2026-03-30"],
            "qr_group": ["ALL", "ALL"],
            "ab_group": ["control", "test_1"],
            "users": [100, 110],
            "arpu": [2.5, 2.7],
        }
    )

    output = tmp_path / "multi_value.xlsx"
    tables = pivot_and_break_table(
        df=df,
        rows="metric",
        value=["users", "arpu"],
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0].to_dict(orient="records") == [
        {"metric": "users", "control": 100.0, "test_1": 110.0},
        {"metric": "arpu", "control": 2.5, "test_1": 2.7},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:4] == [
            ("ALL", None, None),
            ("metric", "control", "test_1"),
            ("users", 100, 110),
            ("arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_detects_value_columns_when_omitted(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "start_dt": ["2026-03-30", "2026-03-30"],
            "qr_group": ["ALL", "ALL"],
            "ab_group": ["control", "test_1"],
            "users": [100, 110],
            "arpu": [2.5, 2.7],
        }
    )

    output = tmp_path / "auto_value.xlsx"
    tables = pivot_and_break_table(
        df=df,
        rows="metric",
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0].to_dict(orient="records") == [
        {"metric": "users", "control": 100.0, "test_1": 110.0},
        {"metric": "arpu", "control": 2.5, "test_1": 2.7},
    ]


def test_pivot_and_break_table_accepts_multiple_dataframes_side_by_side(tmp_path: Path) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [2.5, 2.7],
        }
    )

    output = tmp_path / "multi_df_pivot.xlsx"
    tables = pivot_and_break_table(
        df=[first_df, second_df],
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0][0].to_dict(orient="records") == [
        {"metric": "users", "control": 100, "test_1": 110},
    ]
    assert tables["2026-03-30"][1][0].to_dict(orient="records") == [
        {"metric": "arpu", "control": 2.5, "test_1": 2.7},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:3] == [
            ("ALL", None, None, None, "ALL", None, None),
            ("metric", "control", "test_1", None, "metric", "control", "test_1"),
            ("users", 100, 110, None, "arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_break_table_writes_grouped_raw_tables_without_uniqueness_checks(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users", "users", "users"],
            "ab_group": ["control", "control", "test_1", "test_1"],
            "qr_group": ["ALL", "ALL", "ALL", "ALL"],
            "start_dt": ["2026-03-30"] * 4,
            "value": [100, 120, 110, 130],
        }
    )

    output = tmp_path / "raw_tables.xlsx"
    tables = break_table(
        df=df,
        output=output,
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert list(tables) == ["2026-03-30"]
    assert tables["2026-03-30"][0].to_dict(orient="records") == [
        {"metric": "users", "ab_group": "control", "value": 100},
        {"metric": "users", "ab_group": "control", "value": 120},
        {"metric": "users", "ab_group": "test_1", "value": 110},
        {"metric": "users", "ab_group": "test_1", "value": 130},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-03-30"]
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:6] == [
            ("ALL", None, None),
            ("metric", "ab_group", "value"),
            ("users", "control", 100),
            ("users", "control", 120),
            ("users", "test_1", 110),
            ("users", "test_1", 130),
        ]
    finally:
        workbook.close()


def test_break_table_accepts_multiple_dataframes_side_by_side(tmp_path: Path) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [2.5, 2.7],
        }
    )

    output = tmp_path / "multi_df_raw.xlsx"
    tables = break_table(
        df=[first_df, second_df],
        output=output,
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0][0].to_dict(orient="records") == [
        {"metric": "users", "ab_group": "control", "value": 100},
        {"metric": "users", "ab_group": "test_1", "value": 110},
    ]
    assert tables["2026-03-30"][1][0].to_dict(orient="records") == [
        {"metric": "arpu", "ab_group": "control", "value": 2.5},
        {"metric": "arpu", "ab_group": "test_1", "value": 2.7},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:4] == [
            ("ALL", None, None, None, "ALL", None, None),
            ("metric", "ab_group", "value", None, "metric", "ab_group", "value"),
            ("users", "control", 100, None, "arpu", "control", 2.5),
            ("users", "test_1", 110, None, "arpu", "test_1", 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_replaces_existing_workbook_by_default(tmp_path: Path) -> None:
    output = tmp_path / "replace.xlsx"

    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-04-01", "2026-04-01"],
            "value": [2.5, 2.7],
        }
    )

    pivot_and_break_table(
        df=first_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
    )
    pivot_and_break_table(
        df=second_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-04-01"]
        rows = list(workbook["2026-04-01"].iter_rows(values_only=True))
        assert rows[:2] == [
            ("metric", "control", "test_1"),
            ("arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_appends_new_sheets_when_requested(tmp_path: Path) -> None:
    output = tmp_path / "append.xlsx"

    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-04-01", "2026-04-01"],
            "value": [2.5, 2.7],
        }
    )

    pivot_and_break_table(
        df=first_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
    )
    pivot_and_break_table(
        df=second_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
        append=True,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-03-30", "2026-04-01"]
    finally:
        workbook.close()
